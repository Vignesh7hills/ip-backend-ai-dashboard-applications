"""
Excel generator for TDS Returns — Form 26Q format.

Output sheet 'Annex-I_DeducteeDetails' exactly matches the desired output:
  Col A: Challan Serial No.
  Col B: Section Code
  Col C: Permanent Account Number (PAN) of deductee
  Col D: Name of Deductee
  Col E: Amount of Payment
  Col F: Date on which Amount paid / credited
  Col G: Rate at which Tax deducted
  Col H: Amount of Tax deducted
  Col I: Total Tax Deposited
  Col J: Date on which tax deducted

Also produces a Summary sheet with section-wise totals and
a Challan-wise breakdown.
"""
import io
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.modules.tds_returns.parser import TDSEntry
from app.modules.tds_returns.calculator import ChallanGroup, SectionSummary
from app.core.logger import get_logger

logger = get_logger(__name__)

# Styles
_HDR_FONT  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
_HDR_FILL  = PatternFill(fill_type='solid', fgColor='1F4E79')
_DATA_FONT = Font(name='Arial', size=10)
_ALT_FILL  = PatternFill(fill_type='solid', fgColor='EBF3FB')
_WARN_FILL = PatternFill(fill_type='solid', fgColor='FFCCCC')
_TOTAL_FONT = Font(name='Arial', bold=True, size=10)
_TOTAL_FILL = PatternFill(fill_type='solid', fgColor='FFF2CC')
_THIN       = Side(style='thin', color='AAAAAA')
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_AMT_FMT    = '#,##0.00'

# Exact column headers matching desired output
# Col 11 carries the source bank challan number (reference only — not a Form 26Q field)
_FORM26Q_HEADERS = [
    ('Challan Serial No.',                          8),
    ('Section Code',                               12),
    ('Permanent Account Number (PAN) of deductee', 16),
    ('Name of Deductee',                           40),
    ('Amount of Payment',                          18),
    ('Date on which Amount paid / credited',       22),
    ('Rate at which Tax deducted',                 14),
    ('Amount of Tax deducted',                     18),
    ('Total Tax Deposited',                        18),
    ('Date on which tax deducted',                 22),
]


class TDSReturnsExcelGenerator:

    def generate(
        self,
        challan_groups: List[ChallanGroup],
        company_name: str = '',
        errors: List[str] = None,
        warnings: List[str] = None,
        sections: Dict[str, SectionSummary] = None,
    ) -> bytes:
        wb = openpyxl.Workbook()
        self._build_form26q_sheet(wb, challan_groups, company_name)
        self._build_summary_sheet(wb, challan_groups, sections or {}, company_name)
        if errors or warnings:
            self._build_notes_sheet(wb, errors or [], warnings or [])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Annex-I: Form 26Q deductee details ───────────────────────────────────

    def _build_form26q_sheet(
        self, wb: openpyxl.Workbook,
        challan_groups: List[ChallanGroup],
        company: str
    ):
        ws = wb.active
        ws.title = 'Annex-I_DeducteeDetails'

        ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
        rgt = Alignment(horizontal='right',  vertical='center')
        lft = Alignment(horizontal='left',   vertical='center')

        # Row 1: headers
        for ci, (header, width) in enumerate(_FORM26Q_HEADERS, 1):
            c = ws.cell(row=1, column=ci, value=header)
            c.font = _HDR_FONT
            c.fill = _HDR_FILL
            c.alignment = ctr
            c.border = _BORDER
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 30
        # Col 11: Bank Challan No. (source reference, not a Form 26Q field)
        c11 = ws.cell(row=1, column=11, value='Bank Challan No. (Source)')
        c11.font = _HDR_FONT
        c11.fill = _HDR_FILL
        c11.alignment = ctr
        ws.column_dimensions['K'].width = 20

        r = 2
        alt = False
        for grp in challan_groups:
            for entry in grp.entries:
                fill = _ALT_FILL if alt else None
                inv_pan = not entry.pan_valid and entry.pan

                row_vals = [
                    grp.challan_serial,
                    entry.section,
                    entry.pan,
                    entry.deductee_name,
                    entry.amount_paid,
                    entry.payment_date_str,
                    entry.rate if entry.rate > 0 else None,
                    entry.tds_deducted,
                    entry.tds_deducted,     # Total Tax Deposited = Amount of Tax deducted
                    entry.payment_date_str, # Date tax deducted = same as payment date
                ]
                # Write bank challan no. (from source file) into col 11 if present
                # Col 10 is unused per Form26Q spec; we use it to carry source challan ref.
                bank_ch = getattr(entry, 'bank_challan_no', '')
                if bank_ch:
                    ws.cell(row=r, column=11, value=bank_ch).font = _DATA_FONT

                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=r, column=ci, value=val)
                    c.font = _DATA_FONT
                    c.border = _BORDER

                    if inv_pan and ci == 3:
                        c.fill = _WARN_FILL
                    elif fill:
                        c.fill = fill

                    if ci in (1, 7):
                        c.alignment = ctr
                    elif ci in (5, 8, 9):
                        c.number_format = _AMT_FMT
                        c.alignment = rgt
                    elif ci in (6, 10):
                        c.alignment = ctr
                    else:
                        c.alignment = lft

                ws.row_dimensions[r].height = 16
                r += 1
            alt = not alt

        ws.freeze_panes = 'A2'

    # ── Summary sheet ─────────────────────────────────────────────────────────

    def _build_summary_sheet(
        self, wb, challan_groups, sections, company
    ):
        ws = wb.create_sheet('Summary')
        ctr = Alignment(horizontal='center', vertical='center')
        rgt = Alignment(horizontal='right',  vertical='center')
        lft = Alignment(horizontal='left',   vertical='center')

        r = 1
        ws.merge_cells('A1:G1')
        ws['A1'].value = f"TDS RETURN SUMMARY{' — ' + company if company else ''}"
        ws['A1'].font = Font(name='Arial', bold=True, size=13)
        ws['A1'].alignment = ctr
        ws.row_dimensions[1].height = 24
        r = 3

        # Challan-wise table
        ws.cell(row=r, column=1, value='CHALLAN-WISE BREAKDOWN').font = Font(name='Arial', bold=True, size=11)
        r += 1

        ch_headers = ['Challan No.', 'Section', 'Date', 'No. of Deductees', 'Total Amount', 'Total TDS']
        ch_widths  = [12, 12, 16, 16, 18, 18]
        for ci, (h, w) in enumerate(zip(ch_headers, ch_widths), 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _HDR_FONT
            c.fill = _HDR_FILL
            c.alignment = ctr
            ws.column_dimensions[get_column_letter(ci)].width = w
        r += 1

        alt = False
        for grp in challan_groups:
            fill = _ALT_FILL if alt else None
            for ci, val in enumerate([
                grp.challan_serial, grp.section, grp.payment_date,
                grp.deductee_count, grp.total_amount, grp.total_tds
            ], 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = _DATA_FONT
                c.border = _BORDER
                if fill: c.fill = fill
                if ci in (5, 6):
                    c.number_format = _AMT_FMT
                    c.alignment = rgt
                elif ci in (1, 4):
                    c.alignment = ctr
                else:
                    c.alignment = lft
            r += 1
            alt = not alt

        # Grand total
        tot_amt = sum(g.total_amount for g in challan_groups)
        tot_tds = sum(g.total_tds for g in challan_groups)
        tot_cnt = sum(g.deductee_count for g in challan_groups)
        for ci, val in enumerate(['', 'GRAND TOTAL', '', tot_cnt, tot_amt, tot_tds], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = _TOTAL_FONT
            c.fill = _TOTAL_FILL
            c.border = _BORDER
            if ci in (5, 6):
                c.number_format = _AMT_FMT
                c.alignment = rgt
            elif ci == 4:
                c.alignment = ctr

        r += 2

        # Section-wise table
        ws.cell(row=r, column=1, value='SECTION-WISE SUMMARY').font = Font(name='Arial', bold=True, size=11)
        r += 1

        sec_headers = ['Section', 'Deductees', 'Total Amount', 'Total TDS']
        sec_widths  = [12, 12, 18, 18]
        for ci, (h, w) in enumerate(zip(sec_headers, sec_widths), 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _HDR_FONT
            c.fill = _HDR_FILL
            c.alignment = ctr
            ws.column_dimensions[get_column_letter(ci)].width = w
        r += 1

        # Aggregate per section from challan groups
        sec_agg: Dict[str, dict] = {}
        for grp in challan_groups:
            s = grp.section
            if s not in sec_agg:
                sec_agg[s] = {'cnt': 0, 'amt': 0.0, 'tds': 0.0}
            sec_agg[s]['cnt'] += grp.deductee_count
            sec_agg[s]['amt'] += grp.total_amount
            sec_agg[s]['tds'] += grp.total_tds

        for si, (sec, agg) in enumerate(sorted(sec_agg.items())):
            fill = _ALT_FILL if si % 2 == 1 else None
            for ci, val in enumerate([sec, agg['cnt'], agg['amt'], agg['tds']], 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = _DATA_FONT
                c.border = _BORDER
                if fill: c.fill = fill
                if ci in (3, 4):
                    c.number_format = _AMT_FMT
                    c.alignment = rgt
                elif ci == 2:
                    c.alignment = ctr
                else:
                    c.alignment = lft
            r += 1

        ws.freeze_panes = 'A2'

    # ── Notes sheet ───────────────────────────────────────────────────────────

    def _build_notes_sheet(self, wb, errors, warnings):
        ws = wb.create_sheet('Notes')
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 100
        r = 1
        ws.cell(row=r, column=1, value='Level').font = Font(bold=True)
        ws.cell(row=r, column=2, value='Message').font = Font(bold=True)
        r += 1
        for e in errors:
            ws.cell(row=r, column=1, value='ERROR').font = Font(bold=True, color='9C0006')
            ws.cell(row=r, column=2, value=e)
            r += 1
        for w in warnings:
            ws.cell(row=r, column=1, value='WARNING').font = Font(color='7F4F00')
            ws.cell(row=r, column=2, value=w)
            r += 1
