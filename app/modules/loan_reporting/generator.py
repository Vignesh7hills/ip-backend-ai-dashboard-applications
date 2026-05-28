"""
Excel report generator for Loan Reporting.

Output format matches the client-provided desired output:
  Row 1: "LOAN REPORTING"  (merged, bold, large)
  Row 2: empty
  Row 3: Column headers (Sr No, Name, Address, PAN, Taken, Repaid, Maximum, Squard Up)
  Row 4+: Data rows

Formatting:
  - Header row: bold, blue background, white text
  - Amount columns: Indian number format (comma-separated)
  - YES/NO column: colour-coded (green/red)
  - Auto column widths
"""
import io
from typing import List
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from app.models.account import LoanReportRow
from app.utils.amount_parser import format_indian
from app.core.logger import get_logger

logger = get_logger(__name__)

# Colour palette
_TITLE_FONT = Font(name='Calibri', bold=True, size=14)
_HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
_HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E79')
_DATA_FONT = Font(name='Calibri', size=10)
_ALT_FILL = PatternFill(fill_type='solid', fgColor='EBF3FB')
_YES_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')
_NO_FILL = PatternFill(fill_type='solid', fgColor='FFCCCC')
_YES_FONT = Font(name='Calibri', size=10, bold=True, color='276221')
_NO_FONT = Font(name='Calibri', size=10, bold=True, color='9C0006')
_BORDER_SIDE = Side(style='thin', color='CCCCCC')
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE
)
_AMOUNT_FMT = '#,##0.00'

_COLUMNS = [
    ('Sr No', 7),
    ('Name', 35),
    ('Address', 30),
    ('PAN', 14),
    ('Taken', 18),
    ('Repaid', 18),
    ('Maximum', 22),
    ('Squard Up', 12),
]


class LoanReportingExcelGenerator:

    def generate(self, rows: List[LoanReportRow], company_name: str = '') -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Loan Reporting'

        # ── Row 1: title ──────────────────────────────────────────────────────
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_text = "LOAN REPORTING"
        if company_name:
            title_text = f"LOAN REPORTING - {company_name}"
        title_cell.value = title_text
        title_cell.font = _TITLE_FONT
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # ── Row 2: blank ──────────────────────────────────────────────────────
        ws.row_dimensions[2].height = 6

        # ── Row 3: column headers ─────────────────────────────────────────────
        for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _CELL_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[3].height = 22

        # ── Data rows ─────────────────────────────────────────────────────────
        for i, row in enumerate(rows):
            r = i + 4
            fill = _ALT_FILL if i % 2 == 1 else None
            values = [
                row.sr_no,
                row.name,
                row.address,
                row.pan,
                row.taken,
                row.repaid,
                row.maximum,
                row.squared_up,
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.font = _DATA_FONT
                cell.border = _CELL_BORDER
                if fill:
                    cell.fill = fill

                # Amount columns: right-aligned, formatted
                if col_idx in (5, 6, 7):
                    cell.number_format = _AMOUNT_FMT
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx == 8:
                    # Squared Up
                    if val == 'YES':
                        cell.fill = _YES_FILL
                        cell.font = _YES_FONT
                    else:
                        cell.fill = _NO_FILL
                        cell.font = _NO_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[r].height = 18

        # ── Summary row ───────────────────────────────────────────────────────
        if rows:
            summary_row = len(rows) + 4
            ws.cell(row=summary_row, column=1, value='TOTAL')
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            for col_idx in (5, 6, 7):
                col_letter = get_column_letter(col_idx)
                start = 4
                end = len(rows) + 3
                cell = ws.cell(
                    row=summary_row, column=col_idx,
                    value=f"=SUM({col_letter}{start}:{col_letter}{end})"
                )
                cell.font = Font(bold=True)
                cell.number_format = _AMOUNT_FMT
                cell.alignment = Alignment(horizontal='right')
                cell.border = _CELL_BORDER

        # ── Freeze panes ──────────────────────────────────────────────────────
        ws.freeze_panes = 'A4'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info("Generated Loan Reporting Excel: %d rows", len(rows))
        return buf.read()
