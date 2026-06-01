"""
Excel generator for Trial Balance — Genius Software Import Format.

Output matches the exact structure of the Desired Output / Sample Output File:
  Sheet 1 (TB Import):   Sr.No | Ledger Name | Group | Amount (Dr) | Amount (Cr)
                          Row 1 = header, Row 2 = totals, Row 3+ = entries
  Sheet 2 (Summary):     Trading & P&L summary + Balance Sheet summary
  Sheet 3 (Validation):  Debit total, Credit total, difference, notes
"""

import io
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.modules.trial_balance.parser import TrialBalanceEntry
from app.modules.trial_balance.calculator import GroupSummary
from app.core.logger import get_logger

logger = get_logger(__name__)

# ── Style constants ────────────────────────────────────────────────────────────
_HEADER_FONT  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
_HEADER_FILL  = PatternFill(fill_type='solid', fgColor='1F4E79')
_TOTAL_FONT   = Font(name='Arial', bold=True, size=10)
_TOTAL_FILL   = PatternFill(fill_type='solid', fgColor='FFF2CC')
_DATA_FONT    = Font(name='Arial', size=10)
_SECTION_FONT = Font(name='Arial', bold=True, size=10)
_SECTION_FILL = PatternFill(fill_type='solid', fgColor='D9E1F2')
_THIN          = Side(style='thin', color='AAAAAA')
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_AMT_FMT       = '#,##0.00'
_AMT_FMT_BLANK = '#,##0.00;-#,##0.00;""'   # hide zeros as blank (like desired output)

# Groups that belong to Trading & P&L
_PL_GROUPS = {
    'OPENING STOCK', 'PURCHASE A/C', 'SALES A/C', 'MANUFACTURING EXPENSES',
    'DIRECT EXPENSES', 'DIRECT EXPENSES (M)', 'DIRECT WAGES', 'POWER AND FUEL (M)',
    'CARRIAGE INWARD', 'DIRECT INCOMES', 'CLOSING STOCK',
    # Indirect
    'INDIRECT EXPENSES', 'INDIRECT INCOMES', 'ADMINISTRATIVE EXPENSES',
    'COMMISSION PAID', 'COMMISSION RECEIVED', 'COMPENSATION TO EMPLOYEES',
    'ADVERTISEMENT', 'FREIGHT', 'INSURANCE', 'TELEPHONE', 'TRAVELLING',
    'CONVEYANCE', 'RENT', 'RATES AND TAXES', 'REPAIR & MAINTENANCE',
    'POWER AND FUEL', 'DEPRECIATION', 'BAD DEBTS', 'DONATION',
    'OTHER EXPENSES', 'OTHER INCOMES', 'AUDITORS REMUNERATION',
    'INTEREST PAID', 'INTEREST RECEIVED', 'FINANCIAL EXPENSES',
    'EXTRA-ORDINARY EXPENSES', 'STAFF WELFARE',
}

# Groups that are Credit-normal (Liabilities/Income side)
_CREDIT_NORMAL = {
    'CAPITAL', 'UNSECURED LOANS', 'SECURED LOANS', 'BANK OCC A/C', 'BANK OD A/C',
    'SUNDRY CREDITORS', 'OTHER CREDITORS', 'CREDITORS FOR GOODS', 'CREDITORS FOR EXPENSES',
    'PROVISIONS', 'OUTSTANDING LIABILITIES', 'DEFER TAX LIABILITY',
    'DEFERRED TAX LIABILITIES', 'OTHER CURRENT LIABILITIES',
    'PUBLIC DEPOSITS', 'DEBENTURES', 'DEFFERED PAYMENT CREDIT',
    'SALES A/C', 'DIRECT INCOMES', 'INDIRECT INCOMES', 'INTEREST RECEIVED',
    'COMMISSION RECEIVED', 'RENT INCOME', 'DIVIDEND INCOME', 'OTHER INCOMES',
    'RESERVES AND SURPLUSES', 'SHREE GANESH JI MAHARAJ',
}


def _w(ws, row, col, val, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws.cell(row=row, column=col, value=val)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border
    return c


class TrialBalanceExcelGenerator:

    def generate(
        self,
        groups: Dict[str, GroupSummary],
        entries_flat: List[TrialBalanceEntry] = None,
        company_name: str = '',
        validation_errors: List[str] = None,
        warnings: List[str] = None,
    ) -> bytes:
        wb = openpyxl.Workbook()

        self._build_tb_sheet(wb, groups, company_name)
        self._build_summary_sheet(wb, groups, company_name)
        self._build_validation_sheet(wb, groups, validation_errors or [], warnings or [])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Sheet 1: TB Import (Genius Software format) ────────────────────────────
    def _build_tb_sheet(
        self, wb: openpyxl.Workbook,
        groups: Dict[str, GroupSummary],
        company: str
    ):
        ws = wb.active
        ws.title = 'TB Import'

        # Column widths matching desired output
        for col, w in zip(range(1, 6), [8, 45, 30, 18, 18]):
            ws.column_dimensions[get_column_letter(col)].width = w

        ctr = Alignment(horizontal='center')
        rgt = Alignment(horizontal='right')

        # Row 1 — Header (Sr. No. | Ledger Name | Group | Amount (Dr) | Amount (Cr))
        for ci, h in enumerate(['Sr. No.', 'Ledger Name', 'Group', 'Amount (Dr)', 'Amount (Cr)'], 1):
            _w(ws, 1, ci, h, font=_HEADER_FONT, fill=_HEADER_FILL,
               align=ctr, border=_BORDER)

        # Gather all entries in order (flat)
        all_entries: List[Tuple[TrialBalanceEntry, str]] = []
        for grp in groups.values():
            for e in grp.entries:
                all_entries.append((e, grp.group_name))

        total_dr = sum(e.debit  for e, _ in all_entries)
        total_cr = sum(e.credit for e, _ in all_entries)

        # Row 2 — Totals
        _w(ws, 2, 1, None)
        _w(ws, 2, 2, 'Total', font=_TOTAL_FONT, fill=_TOTAL_FILL)
        _w(ws, 2, 3, None, fill=_TOTAL_FILL)
        _w(ws, 2, 4, total_dr, font=_TOTAL_FONT, fill=_TOTAL_FILL,
           align=rgt, fmt=_AMT_FMT)
        _w(ws, 2, 5, total_cr, font=_TOTAL_FONT, fill=_TOTAL_FILL,
           align=rgt, fmt=_AMT_FMT)

        # Data rows — Sr.No sequential, Ledger Name, Group, Dr, Cr
        sr = 1
        for row_idx, (e, grp_name) in enumerate(all_entries, start=3):
            dr_val = e.debit  if e.debit  > 0 else None
            cr_val = e.credit if e.credit > 0 else None

            _w(ws, row_idx, 1, sr,       font=_DATA_FONT, align=ctr)
            _w(ws, row_idx, 2, e.account_name, font=_DATA_FONT)
            _w(ws, row_idx, 3, grp_name, font=_DATA_FONT)
            _w(ws, row_idx, 4, dr_val,   font=_DATA_FONT, align=rgt,
               fmt=_AMT_FMT if dr_val else None)
            _w(ws, row_idx, 5, cr_val,   font=_DATA_FONT, align=rgt,
               fmt=_AMT_FMT if cr_val else None)
            sr += 1

        ws.freeze_panes = 'A3'

    # ── Sheet 2: Financial Summary ─────────────────────────────────────────────
    def _build_summary_sheet(
        self, wb: openpyxl.Workbook,
        groups: Dict[str, GroupSummary],
        company: str
    ):
        ws = wb.create_sheet('Summary')
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        ctr = Alignment(horizontal='center')
        rgt = Alignment(horizontal='right')
        lft = Alignment(horizontal='left')

        r = 1
        # Title
        ws.merge_cells(f'A{r}:D{r}')
        _w(ws, r, 1, f"FINANCIAL SUMMARY{' — ' + company if company else ''}",
           font=Font(name='Arial', bold=True, size=13),
           align=ctr)
        ws.row_dimensions[r].height = 22
        r += 1; r += 1

        def section_header(title):
            nonlocal r
            ws.merge_cells(f'A{r}:D{r}')
            _w(ws, r, 1, title, font=_SECTION_FONT, fill=_SECTION_FILL, align=lft)
            ws.row_dimensions[r].height = 16
            r += 1

        def sub_header():
            nonlocal r
            for ci, h in enumerate(['Particulars', 'Debit (Dr)', 'Credit (Cr)', 'Net'], 1):
                _w(ws, r, ci, h, font=Font(name='Arial', bold=True, size=9),
                   fill=_HEADER_FILL if ci > 1 else None,
                   align=ctr if ci > 1 else lft)
            r += 1

        def grp_row(label, dr, cr, indent=False, bold=False):
            nonlocal r
            fn = Font(name='Arial', bold=bold, size=10)
            net = dr - cr
            pfx = '  ' if indent else ''
            _w(ws, r, 1, pfx + label, font=fn, align=lft)
            _w(ws, r, 2, dr if dr else None, font=fn, align=rgt, fmt=_AMT_FMT)
            _w(ws, r, 3, cr if cr else None, font=fn, align=rgt, fmt=_AMT_FMT)
            _w(ws, r, 4, net, font=fn, align=rgt, fmt=_AMT_FMT)
            r += 1
            return dr, cr

        def get(gname):
            g = groups.get(gname)
            if g:
                return g.total_debit, g.total_credit
            return 0.0, 0.0

        # ── Trading & P&L ──────────────────────────────────────────────────────
        section_header('TRADING & PROFIT & LOSS ACCOUNT')
        for ci, h in enumerate(['Particulars', 'Debit', 'Credit', 'Net'], 1):
            _w(ws, r, ci, h, font=Font(name='Arial', bold=True, size=9, color='FFFFFF'),
               fill=_HEADER_FILL, align=ctr if ci > 1 else lft)
        r += 1

        # Opening Stock
        dr, cr = get('OPENING STOCK')
        grp_row('Opening Stock', dr, cr)

        # Purchases
        dr, cr = get('PURCHASE A/C')
        grp_row('Purchases (Net)', dr, cr)

        # Manufacturing / Direct Expenses
        mfg_groups = ['MANUFACTURING EXPENSES', 'DIRECT EXPENSES (M)', 'DIRECT WAGES',
                       'POWER AND FUEL (M)', 'CARRIAGE INWARD', 'DIRECT EXPENSES']
        mfg_dr = sum(get(g)[0] for g in mfg_groups)
        mfg_cr = sum(get(g)[1] for g in mfg_groups)
        grp_row('Manufacturing & Direct Expenses', mfg_dr, mfg_cr)

        # Sales
        dr, cr = get('SALES A/C')
        grp_row('Sales (Net)', dr, cr)

        # Gross Profit (Sales + Closing - Opening - Purchases - Mfg)
        sales_dr, sales_cr = get('SALES A/C')
        pur_dr, pur_cr     = get('PURCHASE A/C')
        op_dr, op_cr       = get('OPENING STOCK')
        cl_dr, cl_cr       = get('CLOSING STOCK')
        net_sales    = sales_cr - sales_dr
        net_purchase = pur_dr - pur_cr
        net_mfg      = mfg_dr - mfg_cr
        net_opening  = op_dr - op_cr
        net_closing  = cl_cr - cl_dr
        gross_profit = net_sales + net_closing - net_opening - net_purchase - net_mfg
        r_gp = r
        _w(ws, r, 1, 'Gross Profit / (Loss)', font=Font(name='Arial', bold=True, size=10))
        _w(ws, r, 4, gross_profit, font=Font(name='Arial', bold=True, size=10),
           align=rgt, fmt=_AMT_FMT)
        r += 1; r += 1

        # Indirect Income & Expenses
        ind_inc_groups  = ['INDIRECT INCOMES', 'INTEREST RECEIVED', 'COMMISSION RECEIVED',
                            'RENT INCOME', 'DIVIDEND INCOME', 'OTHER INCOMES', 'DIRECT INCOMES']
        ind_exp_groups  = ['INDIRECT EXPENSES', 'COMMISSION PAID', 'COMPENSATION TO EMPLOYEES',
                            'ADVERTISEMENT', 'FREIGHT', 'INSURANCE', 'TELEPHONE', 'TRAVELLING',
                            'CONVEYANCE', 'RENT', 'RATES AND TAXES', 'REPAIR & MAINTENANCE',
                            'POWER AND FUEL', 'DEPRECIATION', 'BAD DEBTS', 'DONATION',
                            'OTHER EXPENSES', 'AUDITORS REMUNERATION', 'INTEREST PAID',
                            'FINANCIAL EXPENSES', 'EXTRA-ORDINARY EXPENSES', 'STAFF WELFARE']

        ind_inc_dr = sum(get(g)[0] for g in ind_inc_groups)
        ind_inc_cr = sum(get(g)[1] for g in ind_inc_groups)
        ind_exp_dr = sum(get(g)[0] for g in ind_exp_groups)
        ind_exp_cr = sum(get(g)[1] for g in ind_exp_groups)

        grp_row('Indirect Income', ind_inc_dr, ind_inc_cr)
        grp_row('Indirect Expenses', ind_exp_dr, ind_exp_cr)

        net_profit = gross_profit + (ind_inc_cr - ind_inc_dr) - (ind_exp_dr - ind_exp_cr)
        _w(ws, r, 1, 'Net Profit / (Loss)', font=Font(name='Arial', bold=True, size=11))
        _w(ws, r, 4, net_profit, font=Font(name='Arial', bold=True, size=11, color='FF0000' if net_profit < 0 else '006100'),
           align=rgt, fmt=_AMT_FMT)
        r += 1; r += 2

        # ── Balance Sheet ──────────────────────────────────────────────────────
        section_header('BALANCE SHEET SUMMARY')
        for ci, h in enumerate(['Particulars', 'Debit', 'Credit', 'Net'], 1):
            _w(ws, r, ci, h, font=Font(name='Arial', bold=True, size=9, color='FFFFFF'),
               fill=_HEADER_FILL, align=ctr if ci > 1 else lft)
        r += 1

        bs_groups = [
            ('CAPITAL',                  'Capital Account'),
            ('RESERVES AND SURPLUSES',   'Reserves & Surplus'),
            ('SECURED LOANS',            'Secured Loans'),
            ('UNSECURED LOANS',          'Unsecured Loans'),
            ('BANK OCC A/C',             'Bank OCC / OD'),
            ('BANK OD A/C',              'Bank OD'),
            ('SUNDRY CREDITORS',         'Sundry Creditors'),
            ('OTHER CREDITORS',          'Other Creditors'),
            ('PROVISIONS',               'Provisions'),
            ('OUTSTANDING LIABILITIES',  'Outstanding Liabilities'),
            ('OTHER CURRENT LIABILITIES','Other Current Liabilities'),
            ('FIXED ASSETS',             'Fixed Assets'),
            ('INVESTMENTS',              'Investments'),
            ('EQUITY SHARES',            'Equity Shares'),
            ('SHARES',                   'Shares'),
            ('DEPOSITS',                 'Deposits'),
            ('SUNDRY DEBTORS',           'Sundry Debtors'),
            ('LOANS AND ADVANCES (ASSETS)', 'Loans & Advances'),
            ('INVENTORY',                'Inventory'),
            ('CLOSING STOCK',            'Closing Stock'),
            ('CASH AND BANK',            'Cash & Bank'),
            ('CASH IN HAND',             'Cash In Hand'),
            ('OTHER CURRENT ASSETS',     'Other Current Assets'),
            ('BALANCE WITH REVENUE AUTHORITY', 'Balance with Revenue Authority'),
            ('MISC EXPENSES (ASSETS)',   'Misc Expenses (Assets)'),
            ('DEFERRED TAX ASSETS',      'Deferred Tax Assets'),
        ]
        for gkey, glabel in bs_groups:
            dr, cr = get(gkey)
            if dr > 0 or cr > 0:
                grp_row(glabel, dr, cr)

        r += 1
        # Grand totals
        all_dr = sum(g.total_debit  for g in groups.values())
        all_cr = sum(g.total_credit for g in groups.values())
        _w(ws, r, 1, 'GRAND TOTAL (TB)',
           font=Font(name='Arial', bold=True, size=10), fill=_TOTAL_FILL)
        _w(ws, r, 2, all_dr, font=_TOTAL_FONT, fill=_TOTAL_FILL, align=rgt, fmt=_AMT_FMT)
        _w(ws, r, 3, all_cr, font=_TOTAL_FONT, fill=_TOTAL_FILL, align=rgt, fmt=_AMT_FMT)
        _w(ws, r, 4, all_dr - all_cr, font=_TOTAL_FONT, fill=_TOTAL_FILL, align=rgt, fmt=_AMT_FMT)

        ws.freeze_panes = 'A3'

    # ── Sheet 3: Validation ────────────────────────────────────────────────────
    def _build_validation_sheet(
        self, wb: openpyxl.Workbook,
        groups: Dict[str, GroupSummary],
        errors: List[str],
        warnings: List[str]
    ):
        ws = wb.create_sheet('Validation')
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 60

        r = 1
        _w(ws, r, 1, 'VALIDATION & NOTES',
           font=Font(name='Arial', bold=True, size=13))
        r += 2

        total_dr = sum(g.total_debit  for g in groups.values())
        total_cr = sum(g.total_credit for g in groups.values())
        diff     = total_dr - total_cr

        checks = [
            ('Total Debit',   total_dr, 'Sum of all debit entries'),
            ('Total Credit',  total_cr, 'Sum of all credit entries'),
            ('Difference',    diff,     '0 = balanced ✓' if abs(diff) < 0.02 else '⚠ NOT BALANCED'),
        ]
        for label, val, note in checks:
            _w(ws, r, 1, label, font=Font(name='Arial', bold=True, size=10))
            _w(ws, r, 2, val,   font=Font(name='Arial', size=10,
                                          color='006100' if abs(val) < 0.02 and label == 'Difference' else '000000'),
               fmt='#,##0.00')
            _w(ws, r, 3, note,  font=Font(name='Arial', size=10))
            r += 1

        r += 1
        if errors:
            _w(ws, r, 1, 'Errors', font=Font(name='Arial', bold=True, size=10, color='9C0006'))
            r += 1
            for e in errors:
                _w(ws, r, 1, e, font=Font(name='Arial', size=10, color='9C0006'))
                r += 1

        if warnings:
            _w(ws, r, 1, 'Warnings', font=Font(name='Arial', bold=True, size=10, color='7F4F00'))
            r += 1
            for w in warnings:
                _w(ws, r, 1, w, font=Font(name='Arial', size=10, color='7F4F00'))
                r += 1

        r += 1
        _w(ws, r, 1, 'Group Breakdown', font=Font(name='Arial', bold=True, size=10))
        r += 1
        for ci, h in enumerate(['Group', 'Debit', 'Credit', 'Count'], 1):
            _w(ws, r, ci, h, font=_HEADER_FONT, fill=_HEADER_FILL)
        r += 1
        for grp in sorted(groups.values(), key=lambda g: g.group_name):
            _w(ws, r, 1, grp.group_name, font=_DATA_FONT)
            _w(ws, r, 2, grp.total_debit,  font=_DATA_FONT, fmt=_AMT_FMT)
            _w(ws, r, 3, grp.total_credit, font=_DATA_FONT, fmt=_AMT_FMT)
            _w(ws, r, 4, len(grp.entries), font=_DATA_FONT)
            r += 1
